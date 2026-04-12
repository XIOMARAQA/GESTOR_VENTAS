"""
Plantilla e importación Excel de ítems (productos/servicios) por empresa.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from apps.core.excel_template_style import style_help_sheet_title_cell, style_import_sheet_header
from apps.inventario.models import Categoria, Item, Marca, UnidadMedida

SHEET_DATA = "Productos"
SHEET_HELP = "Instrucciones"

# Fila 1: encabezados (orden fijo A–G)
HEADERS = [
    "codigo",
    "nombre",
    "es_servicio",
    "activo",
    "unidad_medida_codigo",
    "categoria_nombre",
    "marca_nombre",
]

HEADER_LABELS = [
    "Código (opcional, único)",
    "Nombre",
    "Es servicio (Sí/No)",
    "Activo (Sí/No)",
    "Código unidad medida (ej. UND)",
    "Categoría (nombre, opcional)",
    "Marca (nombre, opcional)",
]


def _norm_bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("", "0", "no", "false", "f", "n"):
        return False
    if s in ("1", "sí", "si", "yes", "true", "t", "s", "y"):
        return True
    return False


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def build_items_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    for col, (key, label) in enumerate(zip(HEADERS, HEADER_LABELS), start=1):
        ws.cell(row=1, column=col, value=label)
        ws.column_dimensions[get_column_letter(col)].width = min(30, max(15, len(label) + 3))

    style_import_sheet_header(ws, len(HEADERS))

    # Fila de ejemplo (el usuario puede borrarla antes de importar)
    example = ["EJ-001", "Ítem de ejemplo — elimine esta fila", "No", "Sí", "UND", "", ""]
    for col, v in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=v)

    wh = wb.create_sheet(SHEET_HELP)
    wh["A1"] = "Importación de productos y servicios"
    style_help_sheet_title_cell(wh, "A1")
    lines = [
        "1. Complete la hoja «Productos» a partir de la fila 2 (puede borrar la fila de ejemplo).",
        "2. «Código»: si lo deja vacío, se crea un ítem nuevo sin código. Si indica un código que ya existe, se actualiza ese ítem.",
        "3. «Es servicio»: Sí = no descuenta stock en ventas típicas.",
        "4. «Unidad medida»: debe existir en Maestros → Unidades (código interno, ej. UND).",
        "5. «Categoría» y «Marca»: nombre exacto de registros ya creados en su empresa; vacío = sin asignar.",
        "6. Guarde como .xlsx y use «Importar Excel» en la pantalla de productos.",
    ]
    for i, line in enumerate(lines, start=3):
        wh.cell(row=i, column=1, value=line)
    wh.column_dimensions["A"].width = 92

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_map(ws) -> dict[str, int]:
    """Mapa nombre_columna_lógica -> índice 1-based de columna."""
    row1 = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 5)]
    col_by_key: dict[str, int] = {}
    for idx, raw in enumerate(row1, start=1):
        if raw is None:
            continue
        s = re.sub(r"\s+", " ", str(raw).strip().lower())
        # coincidencia flexible con etiquetas o claves
        for i, label in enumerate(HEADER_LABELS):
            if s == label.lower() or s == HEADERS[i]:
                col_by_key[HEADERS[i]] = idx
                break
    # Si no hubo coincidencias, usar orden fijo A–G
    if len(col_by_key) < 3:
        col_by_key = {h: i + 1 for i, h in enumerate(HEADERS)}
    return col_by_key


def _get(row: tuple[Any, ...], col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None:
        return None
    if idx <= len(row):
        return row[idx - 1]
    return None


def import_items_xlsx(content: bytes, empresa_id: int) -> dict[str, Any]:
    """
    Importa filas desde la primera hoja con datos o la hoja «Productos».
    Retorna resumen con creados, actualizados y lista de errores por fila.
    """
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb[wb.sheetnames[0]]
    col_map = _header_map(ws)

    creados = 0
    actualizados = 0
    errores: list[dict[str, Any]] = []

    rows_iter = ws.iter_rows(min_row=2, values_only=True)

    with transaction.atomic():
        for row_num, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            row = tuple(row) + tuple()
            nombre = _cell_str(_get(row, col_map, "nombre"))
            if not nombre:
                continue
            # Saltar fila de plantilla de ejemplo
            if nombre.lower().startswith("ítem de ejemplo") or "elimine" in nombre.lower():
                continue

            codigo = _cell_str(_get(row, col_map, "codigo"))
            es_serv = _norm_bool(_get(row, col_map, "es_servicio"))
            raw_act = _get(row, col_map, "activo")
            if raw_act is None or (isinstance(raw_act, str) and not str(raw_act).strip()):
                activo = True
            else:
                activo = _norm_bool(raw_act)
            um_cod = _cell_str(_get(row, col_map, "unidad_medida_codigo")).upper() or "UND"
            cat_n = _cell_str(_get(row, col_map, "categoria_nombre"))
            mar_n = _cell_str(_get(row, col_map, "marca_nombre"))

            try:
                um = UnidadMedida.objects.filter(
                    empresa_id=empresa_id, codigo__iexact=um_cod
                ).first()
                if um is None:
                    raise ValueError(
                        f'Unidad de medida «{um_cod}» no existe. Créela en Maestros → Unidades.'
                    )

                categoria = None
                if cat_n:
                    categoria = Categoria.objects.filter(
                        empresa_id=empresa_id, nombre__iexact=cat_n
                    ).first()
                    if categoria is None:
                        raise ValueError(f'Categoría «{cat_n}» no encontrada.')

                marca = None
                if mar_n:
                    marca = Marca.objects.filter(
                        empresa_id=empresa_id, nombre__iexact=mar_n
                    ).first()
                    if marca is None:
                        raise ValueError(f'Marca «{mar_n}» no encontrada.')

                item = None
                if codigo:
                    item = Item.objects.filter(
                        empresa_id=empresa_id, codigo=codigo
                    ).first()

                if item:
                    item.nombre = nombre[:255]
                    item.es_servicio = es_serv
                    item.activo = activo
                    item.unidad_medida = um
                    item.categoria = categoria
                    item.marca = marca
                    item.save(
                        update_fields=[
                            "nombre",
                            "es_servicio",
                            "activo",
                            "unidad_medida",
                            "categoria",
                            "marca",
                            "actualizado_en",
                        ]
                    )
                    actualizados += 1
                else:
                    Item.objects.create(
                        empresa_id=empresa_id,
                        codigo=codigo[:50] if codigo else "",
                        nombre=nombre[:255],
                        es_servicio=es_serv,
                        activo=activo,
                        unidad_medida=um,
                        categoria=categoria,
                        marca=marca,
                    )
                    creados += 1
            except Exception as e:
                errores.append({"fila": row_num, "mensaje": str(e)})

    return {
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
    }
