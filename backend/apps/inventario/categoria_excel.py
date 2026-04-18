"""
Plantilla e importación Excel de categorías de producto por empresa.
La columna «categoría padre» es el nombre de la categoría superior (opcional);
debe existir ya en base o aparecer antes en el Excel para poder enlazar.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from apps.core.excel_template_style import style_help_sheet_title_cell, style_import_sheet_header
from apps.inventario.models import Categoria

SHEET_DATA = "Categorias"
SHEET_HELP = "Instrucciones"

HEADERS = ["nombre", "categoria_padre", "activo"]

HEADER_LABELS = [
    "Nombre (obligatorio)",
    "Categoría padre (nombre, opcional; vacío = raíz)",
    "Activo (Sí/No)",
]


def _norm_bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("", "0", "no", "false", "f", "n"):
        return False
    if s in ("1", "sí", "si", "yes", "true", "t", "s", "y"):
        return True
    return False


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def build_categorias_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    for col, label in enumerate(HEADER_LABELS, start=1):
        ws.cell(row=1, column=col, value=label)
        ws.column_dimensions[get_column_letter(col)].width = min(48, max(18, len(label) + 3))

    style_import_sheet_header(ws, len(HEADERS))

    example = ["Bebidas", "", "Sí"]
    for col, v in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=v)
    ws.cell(row=3, column=1, value="Gaseosas")
    ws.cell(row=3, column=2, value="Bebidas")
    ws.cell(row=3, column=3, value="Sí")

    wh = wb.create_sheet(SHEET_HELP)
    wh["A1"] = "Importación de categorías"
    style_help_sheet_title_cell(wh, "A1")
    lines = [
        "1. Hoja «Categorias»: datos desde la fila 2 (puede borrar filas de ejemplo).",
        "2. «Categoría padre» debe coincidir con el nombre de otra categoría (sin importar mayúsculas).",
        "3. Si hay varias categorías con el mismo nombre, se usa la primera encontrada como padre.",
        "4. Ordene o repita importación: si el padre aún no existe, se intentará en rondas siguientes.",
        "5. Si tras varias rondas falta el padre, se reportará error en esa fila.",
        "6. Guarde como .xlsx y use «Importar Excel» en la pantalla de categorías.",
    ]
    for i, line in enumerate(lines, start=3):
        wh.cell(row=i, column=1, value=line)
    wh.column_dimensions["A"].width = 92

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_map(ws) -> dict[str, int]:
    row1 = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 3)]
    col_by_key: dict[str, int] = {}
    for idx, raw in enumerate(row1, start=1):
        if raw is None:
            continue
        s = re.sub(r"\s+", " ", str(raw).strip().lower())
        for i, label in enumerate(HEADER_LABELS):
            if s == label.lower() or s == HEADERS[i]:
                col_by_key[HEADERS[i]] = idx
                break
    if "nombre" not in col_by_key:
        col_by_key = {h: i + 1 for i, h in enumerate(HEADERS)}
    return col_by_key


def _get(row: tuple[Any, ...], col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None:
        return None
    if idx <= len(row):
        return row[idx - 1]
    return None


def _find_parent_id(empresa_id: int, padre_nombre: str) -> int | None:
    pn = padre_nombre.strip()
    if not pn:
        return None
    p = (
        Categoria.objects.filter(empresa_id=empresa_id)
        .filter(nombre__iexact=pn)
        .order_by("id")
        .first()
    )
    return p.pk if p else None


def _find_cat(empresa_id: int, nombre: str, padre_id: int | None) -> Categoria | None:
    return (
        Categoria.objects.filter(empresa_id=empresa_id)
        .filter(nombre__iexact=nombre.strip(), padre_id=padre_id)
        .order_by("id")
        .first()
    )


def import_categorias_xlsx(content: bytes, empresa_id: int) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb[wb.sheetnames[0]]
    col_map = _header_map(ws)

    creados = 0
    actualizados = 0
    errores: list[dict[str, Any]] = []

    parsed: list[dict[str, Any]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        row = tuple(row) + tuple()
        nombre = _cell_str(_get(row, col_map, "nombre"))
        if not nombre:
            continue
        low = nombre.lower()
        if "elimine esta fila" in low:
            continue
        padre = _cell_str(_get(row, col_map, "categoria_padre"))
        raw_act = _get(row, col_map, "activo")
        if raw_act is None or (isinstance(raw_act, str) and not str(raw_act).strip()):
            activo = True
        else:
            activo = _norm_bool(raw_act)
        parsed.append(
            {
                "fila": row_num,
                "nombre": nombre[:120],
                "padre": padre[:120] if padre else "",
                "activo": activo,
            }
        )

    max_rounds = max(30, len(parsed) * 3 + 5)

    with transaction.atomic():
        pending = parsed[:]
        for _ in range(max_rounds):
            if not pending:
                break
            nxt: list[dict[str, Any]] = []
            progressed = False
            for rec in pending:
                padre_nom = rec["padre"]
                padre_id: int | None = None
                if padre_nom:
                    if rec["nombre"].lower() == padre_nom.lower():
                        errores.append(
                            {
                                "fila": rec["fila"],
                                "mensaje": "La categoría no puede ser padre de sí misma.",
                            }
                        )
                        progressed = True
                        continue
                    padre_id = _find_parent_id(empresa_id, padre_nom)
                    if padre_id is None:
                        nxt.append(rec)
                        continue

                ex = _find_cat(empresa_id, rec["nombre"], padre_id)
                try:
                    if ex:
                        ex.activo = rec["activo"]
                        ex.save(update_fields=["activo"])
                        actualizados += 1
                    else:
                        Categoria.objects.create(
                            empresa_id=empresa_id,
                            nombre=rec["nombre"],
                            padre_id=padre_id,
                            activo=rec["activo"],
                        )
                        creados += 1
                    progressed = True
                except Exception as e:
                    errores.append({"fila": rec["fila"], "mensaje": str(e)})
                    progressed = True
            pending = nxt
            if not progressed and pending:
                for rec in pending:
                    errores.append(
                        {
                            "fila": rec["fila"],
                            "mensaje": f"No se encontró la categoría padre «{rec['padre']}».",
                        }
                    )
                break

    return {"creados": creados, "actualizados": actualizados, "errores": errores}
