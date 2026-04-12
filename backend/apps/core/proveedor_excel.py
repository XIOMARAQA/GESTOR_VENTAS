"""
Plantilla e importación Excel de proveedores por empresa.
Tabla ``proveedor``: razon_social, documento (opcional), activo.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from apps.core.excel_template_style import style_help_sheet_title_cell, style_import_sheet_header
from apps.core.models import Proveedor

SHEET_DATA = "Proveedores"
SHEET_HELP = "Instrucciones"

HEADERS = [
    "documento",
    "razon_social",
    "activo",
]

HEADER_LABELS = [
    "Documento (RUC/DNI, opcional; clave si ya existe)",
    "Razón social o nombre (obligatorio)",
    "Activo (Sí/No)",
]


def _norm_bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("", "0", "no", "false", "f", "n"):
        return False
    if s in ("1", "sí", "si", "yes", "true", "t", "s", "y"):
        return True
    return False


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def build_proveedores_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    for col, label in enumerate(HEADER_LABELS, start=1):
        ws.cell(row=1, column=col, value=label)
        ws.column_dimensions[get_column_letter(col)].width = min(42, max(18, len(label) + 3))

    style_import_sheet_header(ws, len(HEADERS))

    example = ["20123456789", "Proveedor de ejemplo — elimine esta fila", "Sí"]
    for col, v in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=v)

    wh = wb.create_sheet(SHEET_HELP)
    wh["A1"] = "Importación de proveedores"
    style_help_sheet_title_cell(wh, "A1")
    lines = [
        "1. Hoja «Proveedores»: datos desde la fila 2 (puede borrar la fila de ejemplo).",
        "2. Columnas del maestro: solo documento, razón social y activo (como en la tabla proveedor).",
        "3. Si indica un documento que ya existe en su empresa, se actualiza razón social y activo.",
        "4. Si el documento está vacío, siempre se crea un proveedor nuevo (puede haber duplicados sin documento).",
        "5. Guarde como .xlsx y use «Importar Excel» en la pantalla de proveedores.",
    ]
    for i, line in enumerate(lines, start=3):
        wh.cell(row=i, column=1, value=line)
    wh.column_dimensions["A"].width = 88

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_map(ws) -> dict[str, int]:
    row1 = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 3)]
    col_by_key: dict[str, int] = {}
    for idx, raw in enumerate(row1, start=1):
        if raw is None:
            continue
        s = re.sub(r"\s+", " ", str(raw).strip().lower())
        for i, label in enumerate(HEADER_LABELS):
            if s == label.lower() or s == HEADERS[i]:
                col_by_key[HEADERS[i]] = idx
                break
    if len(col_by_key) < 2:
        col_by_key = {h: i + 1 for i, h in enumerate(HEADERS)}
    return col_by_key


def _get(row: tuple[Any, ...], col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None:
        return None
    if idx <= len(row):
        return row[idx - 1]
    return None


def import_proveedores_xlsx(content: bytes, empresa_id: int) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb[wb.sheetnames[0]]
    col_map = _header_map(ws)

    creados = 0
    actualizados = 0
    errores: list[dict[str, Any]] = []

    rows_iter = ws.iter_rows(min_row=2, values_only=True)

    with transaction.atomic():
        for row_num, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            row = tuple(row) + tuple()
            doc = _cell_str(_get(row, col_map, "documento"))
            razon = _cell_str(_get(row, col_map, "razon_social"))
            if not razon and not doc:
                continue
            low = razon.lower()
            if "elimine esta fila" in low or low.startswith("proveedor de ejemplo"):
                continue
            if not razon:
                errores.append(
                    {"fila": row_num, "mensaje": "Razón social / nombre es obligatorio."}
                )
                continue

            razon = razon[:255]
            doc = doc[:20]
            raw_act = _get(row, col_map, "activo")
            if raw_act is None or (isinstance(raw_act, str) and not str(raw_act).strip()):
                activo = True
            else:
                activo = _norm_bool(raw_act)

            try:
                prov = None
                if doc:
                    prov = Proveedor.objects.filter(
                        empresa_id=empresa_id, documento=doc
                    ).first()
                if prov:
                    prov.razon_social = razon
                    prov.activo = activo
                    prov.save(update_fields=["razon_social", "activo"])
                    actualizados += 1
                else:
                    Proveedor.objects.create(
                        empresa_id=empresa_id,
                        documento=doc,
                        razon_social=razon,
                        activo=activo,
                    )
                    creados += 1
            except Exception as e:
                errores.append({"fila": row_num, "mensaje": str(e)})

    return {"creados": creados, "actualizados": actualizados, "errores": errores}
