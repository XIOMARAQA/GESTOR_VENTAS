"""
Estilos visuales comunes para plantillas Excel de importación (encabezados).
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta alineada al panel (teal / slate)
_HEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_SIDE = Side(style="thin", color="64748B")
_HEADER_BORDER = Border(
    left=_HEADER_SIDE,
    right=_HEADER_SIDE,
    top=_HEADER_SIDE,
    bottom=_HEADER_SIDE,
)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_import_sheet_header(ws, num_cols: int, *, row_height: float = 30) -> None:
    """
    Fila 1: fondo teal, texto blanco, bordes, centrado, altura, filtro y paneles fijos.
    """
    ws.row_dimensions[1].height = row_height
    last_letter = get_column_letter(num_cols)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_letter}1"


def style_help_sheet_title_cell(wh, cell_coord: str = "A1") -> None:
    """Título de la hoja de instrucciones (una celda)."""
    cell = wh[cell_coord]
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.border = _HEADER_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
