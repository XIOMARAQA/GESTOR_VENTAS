"""
Plantilla e importación Excel de vendedores por empresa.
Clave de actualización: DNI de 8 dígitos si viene informado; si no, apellidos y nombres (coincidencia exacta sin mayúsculas).
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from apps.core.excel_template_style import style_help_sheet_title_cell, style_import_sheet_header
from apps.core.models import Sucursal, Vendedor

SHEET_DATA = "Vendedores"
SHEET_HELP = "Instrucciones"

HEADERS = [
    "dni",
    "apellido_paterno",
    "apellido_materno",
    "nombres",
    "sucursal",
    "activo",
]

HEADER_LABELS = [
    "DNI (8 dígitos, opcional; clave preferente si existe)",
    "Apellido paterno",
    "Apellido materno",
    "Nombres",
    "Sucursal (nombre, opcional; debe existir en la empresa)",
    "Activo (Sí/No)",
]


def _norm_bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("", "0", "no", "false", "f", "n"):
        return False
    if s in ("1", "sí", "si", "yes", "true", "t", "s", "y"):
        return True
    return False


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def build_vendedores_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    for col, label in enumerate(HEADER_LABELS, start=1):
        ws.cell(row=1, column=col, value=label)
        ws.column_dimensions[get_column_letter(col)].width = min(42, max(16, len(label) + 2))

    style_import_sheet_header(ws, len(HEADERS))

    example = ["87654321", "Ejemplo", "Apellido", "Nombre", "", "Sí"]
    for col, v in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=v)

    wh = wb.create_sheet(SHEET_HELP)
    wh["A1"] = "Importación de vendedores"
    style_help_sheet_title_cell(wh, "A1")
    lines = [
        "1. Hoja «Vendedores»: datos desde la fila 2 (puede borrar la fila de ejemplo).",
        "2. Si indica DNI de 8 dígitos y ya existe un vendedor con ese DNI en la empresa, se actualizan los demás campos.",
        "3. Sin DNI (o DNI inválido): se busca por apellido paterno, materno y nombres exactos (sin importar mayúsculas).",
        "4. Si no hay coincidencia, se crea un vendedor nuevo.",
        "5. Sucursal: nombre tal como figura en el maestro de sucursales de la misma empresa; si no coincide, se deja vacío.",
        "6. Guarde como .xlsx y use «Importar Excel» en la pantalla de vendedores.",
    ]
    for i, line in enumerate(lines, start=3):
        wh.cell(row=i, column=1, value=line)
    wh.column_dimensions["A"].width = 92

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_map(ws) -> dict[str, int]:
    row1 = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 3)]
    col_by_key: dict[str, int] = {}
    for idx, raw in enumerate(row1, start=1):
        if raw is None:
            continue
        s = re.sub(r"\s+", " ", str(raw).strip().lower())
        for i, label in enumerate(HEADER_LABELS):
            if s == label.lower() or s == HEADERS[i]:
                col_by_key[HEADERS[i]] = idx
                break
    if len(col_by_key) < 3:
        col_by_key = {h: i + 1 for i, h in enumerate(HEADERS)}
    return col_by_key


def _get(row: tuple[Any, ...], col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None:
        return None
    if idx <= len(row):
        return row[idx - 1]
    return None


def _resolve_sucursal_id(empresa_id: int, nombre: str) -> int | None:
    n = nombre.strip()
    if not n:
        return None
    s = Sucursal.objects.filter(empresa_id=empresa_id, nombre__iexact=n).order_by("id").first()
    return s.pk if s else None


def _find_vendedor(empresa_id: int, dni_store: str, ap1: str, ap2: str, nom: str) -> Vendedor | None:
    digits = re.sub(r"\D", "", dni_store)
    if len(digits) == 8:
        return Vendedor.objects.filter(empresa_id=empresa_id, dni=digits).first()
    return (
        Vendedor.objects.filter(empresa_id=empresa_id)
        .filter(
            apellido_paterno__iexact=ap1[:80],
            apellido_materno__iexact=ap2[:80],
            nombres__iexact=nom[:120],
        )
        .order_by("id")
        .first()
    )


def import_vendedores_xlsx(content: bytes, empresa_id: int) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb[wb.sheetnames[0]]
    col_map = _header_map(ws)

    creados = 0
    actualizados = 0
    errores: list[dict[str, Any]] = []

    rows_iter = ws.iter_rows(min_row=2, values_only=True)

    with transaction.atomic():
        for row_num, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            row = tuple(row) + tuple()
            dni_raw = _cell_str(_get(row, col_map, "dni"))
            ap1 = _cell_str(_get(row, col_map, "apellido_paterno"))
            ap2 = _cell_str(_get(row, col_map, "apellido_materno"))
            nom = _cell_str(_get(row, col_map, "nombres"))
            suc_nom = _cell_str(_get(row, col_map, "sucursal"))

            if not dni_raw and not ap1 and not nom:
                continue
            low = (ap1 + nom).lower()
            if "ejemplo" in low and "apellido" in low:
                continue

            dni_digits = re.sub(r"\D", "", dni_raw)[:20]
            if len(dni_digits) == 8:
                dni_store = dni_digits
            elif dni_raw:
                dni_store = dni_raw[:20]
            else:
                dni_store = ""

            raw_act = _get(row, col_map, "activo")
            if raw_act is None or (isinstance(raw_act, str) and not str(raw_act).strip()):
                activo = True
            else:
                activo = _norm_bool(raw_act)

            if not ap1.strip() and not nom.strip() and not dni_store:
                errores.append(
                    {"fila": row_num, "mensaje": "Indique DNI o al menos apellido paterno o nombres."}
                )
                continue

            suc_id = _resolve_sucursal_id(empresa_id, suc_nom)
            if suc_nom and suc_id is None:
                errores.append(
                    {
                        "fila": row_num,
                        "mensaje": f"Sucursal «{suc_nom}» no encontrada para la empresa (se importó sin sucursal).",
                    }
                )

            try:
                v = _find_vendedor(empresa_id, dni_store, ap1, ap2, nom)
                if v:
                    v.dni = dni_store
                    v.apellido_paterno = ap1[:80]
                    v.apellido_materno = ap2[:80]
                    v.nombres = nom[:120]
                    v.sucursal_id = suc_id
                    v.activo = activo
                    v.save(
                        update_fields=[
                            "dni",
                            "apellido_paterno",
                            "apellido_materno",
                            "nombres",
                            "sucursal_id",
                            "activo",
                        ]
                    )
                    actualizados += 1
                else:
                    Vendedor.objects.create(
                        empresa_id=empresa_id,
                        dni=dni_store,
                        apellido_paterno=ap1[:80],
                        apellido_materno=ap2[:80],
                        nombres=nom[:120],
                        sucursal_id=suc_id,
                        activo=activo,
                    )
                    creados += 1
            except Exception as e:
                errores.append({"fila": row_num, "mensaje": str(e)})

    return {"creados": creados, "actualizados": actualizados, "errores": errores}
